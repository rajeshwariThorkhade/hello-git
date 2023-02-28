import calendar
import openpyxl

excel_file = "LicenseDashboard.xlsx"
data = {'Feb': [455, 900, 800, 600, 3989, 655, 9003, 786]}
out_excel = 'test2.xlsx'


def append_data(excel_file, data, out_excel):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    dic = {}
    i = 4
    for num in range(1, 13):
        mon = calendar.month_abbr[num]
        dic[mon] = i
        i = i + 1
    print(ws['D3'].value)
    for key, col in dic.items():
        print(key, col)
        wb = openpyxl.load_workbook(out_excel)
        wss = wb.active
        for k, v in data.items():
            if key == k:
                for i, value in enumerate(data[key], start=3):
                    wss.cell(row=i, column=col).value = value
                    wb.save(out_excel)
                    print("data is appended to month column")


append_data(excel_file, data,  out_excel)
